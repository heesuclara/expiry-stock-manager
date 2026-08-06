import streamlit as st
import pandas as pd
import pyxlsb
import openpyxl
import warnings
import json
import os
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
from collections import defaultdict
from io import BytesIO

warnings.filterwarnings("ignore")

_DIR = os.path.dirname(os.path.abspath(__file__))
HISTORY_PATH = os.path.join(_DIR, "data", "history.json")
RESALE_PATH  = os.path.join(_DIR, "data", "resale_prevention.json")
STATUS_PATH  = os.path.join(_DIR, "data", "product_status.json")

st.set_page_config(
    page_title="유통기한 임박재고 분석",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .block-container { padding-top: 1.5rem; }
    .urgent-badge {
        background: #ff4b4b22; border-left: 4px solid #ff4b4b;
        border-radius: 4px; padding: 0.5rem 1rem; margin-bottom: 0.5rem;
    }
    .warning-badge {
        background: #ffa50022; border-left: 4px solid #ffa500;
        border-radius: 4px; padding: 0.5rem 1rem; margin-bottom: 0.5rem;
    }
    .orange-badge {
        background: #ff8c0022; border-left: 4px solid #ff8c00;
        border-radius: 4px; padding: 0.5rem 1rem; margin-bottom: 0.5rem;
    }
    .ok-badge {
        background: #21c35422; border-left: 4px solid #21c354;
        border-radius: 4px; padding: 0.5rem 1rem; margin-bottom: 0.5rem;
    }
    .resale-badge {
        background: #e879091a; border-left: 3px solid #e87909;
        border-radius: 4px; padding: 0.4rem 0.9rem;
        font-size: 0.88em; margin-top: 0.4rem;
    }
    .incoming-badge {
        background: #3b82f61a; border-left: 3px solid #3b82f6;
        border-radius: 4px; padding: 0.4rem 0.9rem;
        font-size: 0.88em; margin-top: 0.4rem;
    }
    .nostock-badge {
        background: #ef44441a; border-left: 3px solid #ef4444;
        border-radius: 4px; padding: 0.4rem 0.9rem;
        font-size: 0.88em; margin-top: 0.4rem;
    }
    .f2-note {
        background: #7c3aed11; border-left: 3px solid #7c3aed;
        border-radius: 4px; padding: 0.3rem 0.8rem;
        font-size: 0.85em; color: #7c3aed; margin-top: 0.3rem;
    }
    div[data-testid="stExpander"] { border-radius: 8px; margin-bottom: 6px; }
</style>
""", unsafe_allow_html=True)


# ─── 헬퍼 함수 ──────────────────────────────────────────────────
def excel_serial_to_date(n):
    if n is None:
        return None
    if isinstance(n, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=int(n))).date()
        except Exception:
            return None
    if isinstance(n, str):
        try:
            return datetime.strptime(n.strip(), "%Y-%m-%d").date()
        except Exception:
            return None
    return None


def months_remaining(target, base):
    if target is None or base is None:
        return None
    rd = relativedelta(target, base)
    return rd.years * 12 + rd.months


def parse_category(raw_text):
    if not raw_text or not isinstance(raw_text, str):
        return None
    first_line = raw_text.split("\n")[0].strip()
    if "7" in first_line or "9" in first_line:
        return "7-9개월"
    if "4" in first_line or "6" in first_line:
        return "4-6개월"
    if "3" in first_line:
        return "3개월"
    return first_line


def adjust_to_weekday(d):
    if d.weekday() == 5:
        return d - timedelta(days=1)
    elif d.weekday() == 6:
        return d - timedelta(days=2)
    return d


# ─── 재고수불종합 시트 로더 ───────────────────────────────────────
@st.cache_data(show_spinner=False)
def read_stock_sheet(file_bytes: bytes) -> dict:
    """재고수불종합 → 담당팀 / 브랜드(스타일코드) / 상품코드 매핑 반환"""
    result: dict = {
        "teams": [],            # 담당팀 목록 (순서 유지)
        "team_brands": {},      # {팀: [브랜드, ...]}
        "team_brand_codes": {}, # {팀: {브랜드: [코드, ...]}}
        "overseas_codes": set(),# 품목구분 = "해외전용" 상품코드
        "code_to_brand": {},    # {코드: 브랜드}
    }

    try:
        with pyxlsb.open_workbook(BytesIO(file_bytes)) as wb:
            if "재고수불종합" not in wb.sheets:
                return result
            with wb.get_sheet("재고수불종합") as sheet:
                for i, row in enumerate(sheet.rows()):
                    if i < 4:   # 헤더: 인덱스3, 데이터: 인덱스4부터
                        continue
                    vals = [c.v for c in row]
                    if len(vals) < 6:
                        continue

                    team       = vals[1]   # 담당팀
                    품목구분   = vals[2]   # C / 해외전용
                    style_code = vals[4]   # 스타일코드 = 브랜드명
                    code       = vals[5]   # 상품코드

                    if not team or not isinstance(team, str):
                        continue
                    if not code:
                        continue
                    if not style_code or not isinstance(style_code, str):
                        continue

                    team       = team.strip()
                    style_code = style_code.strip()
                    code       = str(code).strip()
                    품목구분_s = 품목구분.strip() if isinstance(품목구분, str) else ""

                    if 품목구분_s == "해외전용":
                        result["overseas_codes"].add(code)

                    if team not in result["team_brands"]:
                        result["teams"].append(team)
                        result["team_brands"][team] = []
                        result["team_brand_codes"][team] = {}

                    if style_code not in result["team_brands"][team]:
                        result["team_brands"][team].append(style_code)
                        result["team_brand_codes"][team][style_code] = []

                    if code not in result["team_brand_codes"][team][style_code]:
                        result["team_brand_codes"][team][style_code].append(code)

                    result["code_to_brand"][code] = style_code
    except Exception:
        pass

    return result


# ─── 입고예정(구매) 로더 ──────────────────────────────────────────
@st.cache_data(show_spinner=False)
def read_incoming_sheet(file_bytes: bytes) -> dict:
    """입고예정(구매) 시트 → {상품코드: {잔량, 현재입고예정일}} 반환"""
    _BASE = datetime(1899, 12, 30)
    def _to_str(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        try:
            return (_BASE + timedelta(days=int(v))).strftime("%Y-%m-%d")
        except Exception:
            return str(v)

    result: dict = {}
    try:
        with pyxlsb.open_workbook(BytesIO(file_bytes)) as wb:
            with wb.get_sheet("입고예정(구매)") as sheet:
                rows = list(sheet.rows())
                for row in rows[3:]:          # header=row2, data from row3
                    vals = [c.v for c in row]
                    if len(vals) <= 3:
                        continue
                    code = vals[3]
                    if not isinstance(code, str) or not code.strip():
                        continue
                    code = code.strip()
                    잔량_raw = vals[10] if len(vals) > 10 else 0
                    예정일_raw = vals[12] if len(vals) > 12 else None
                    try:
                        잔량 = int(잔량_raw or 0)
                    except (ValueError, TypeError):
                        잔량 = 0
                    예정일 = _to_str(예정일_raw)

                    if code not in result:
                        result[code] = {"잔량": 0, "현재입고예정일": ""}
                    result[code]["잔량"] += 잔량
                    if not result[code]["현재입고예정일"] and 예정일:
                        result[code]["현재입고예정일"] = 예정일
    except Exception:
        pass
    return result


# ─── 유통기한관리 로더 ───────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_file1(file_bytes: bytes, code_whitelist: tuple, code_to_brand_tuple: tuple) -> dict:
    """유통기한관리 시트 → 선택된 상품코드의 로트 정보"""
    code_to_brand = dict(code_to_brand_tuple)
    whitelist_set = set(code_whitelist)

    brand_products: dict = {}
    with pyxlsb.open_workbook(BytesIO(file_bytes)) as wb:
        with wb.get_sheet("유통기한관리") as sheet:
            for i, row in enumerate(sheet.rows()):
                if i < 5:
                    continue
                vals = [c.v for c in row]
                if len(vals) < 15:
                    continue

                코드 = vals[7]
                if not 코드:
                    continue
                코드 = str(코드).strip()

                if 코드 not in whitelist_set:
                    continue

                품명 = vals[8]
                if not 품명 or not isinstance(품명, str):
                    continue

                brand = code_to_brand.get(코드, 품명.split()[0] if 품명 else "기타")
                유통기한 = excel_serial_to_date(vals[13])
                재고     = vals[14]

                if 코드 not in brand_products:
                    brand_products[코드] = {
                        "name": 품명,
                        "brand": brand,
                        "lots": [],
                    }
                stock_qty = int(재고) if 재고 else 0
                if 유통기한 is not None and stock_qty > 0:
                    brand_products[코드]["lots"].append({
                        "expiry": 유통기한,
                        "stock": stock_qty,
                    })
    return brand_products


@st.cache_data(show_spinner=False)
def load_file2(file_bytes: bytes) -> tuple:
    """임박재고 xlsx → (카테고리 dict, 소비기한 dict)"""
    product_cats: dict = {}
    product_f2_dates: dict = defaultdict(list)

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
    # "상품별수불현황"이 포함된 시트 이름을 우선 탐색 (예: _전사 공유용 등 접미사 변경 대응)
    ws = None
    for _name in wb.sheetnames:
        if "상품별수불현황" in _name:
            ws = wb[_name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]   # 없으면 첫 번째 시트
    current_cat = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        구분, 상품코드, _, 소비기한 = row[1], row[2], row[3], row[4]
        if 구분 and isinstance(구분, str) and 구분.strip():
            current_cat = parse_category(구분)
        if 상품코드:
            if 상품코드 not in product_cats:
                product_cats[상품코드] = current_cat
            if 소비기한:
                try:
                    d = datetime.strptime(str(소비기한), "%Y-%m-%d").date()
                    product_f2_dates[상품코드].append(d)
                except Exception:
                    pass
    wb.close()
    return product_cats, dict(product_f2_dates)


# ─── 재판매방지 로더 ─────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_resale_prevention() -> dict:
    """data/resale_prevention.json → {상품코드: 재판매방지 정보}"""
    if os.path.exists(RESALE_PATH):
        try:
            with open(RESALE_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


# ─── 품절/단종 상태 관리 ─────────────────────────────────────────
def load_product_status() -> list:
    """data/product_status.json → 품절/단종 키워드 목록"""
    if os.path.exists(STATUS_PATH):
        try:
            with open(STATUS_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return []

def save_product_status(data: list):
    """품절/단종 목록 저장"""
    os.makedirs(os.path.dirname(STATUS_PATH), exist_ok=True)
    with open(STATUS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def match_product_status(name: str, status_list: list):
    """품명에 키워드가 포함되면 (status, keyword) 반환, 없으면 (None, None)
    더 긴 키워드(구체적인 것)를 우선 매칭"""
    for item in sorted(status_list, key=lambda x: len(x["keyword"]), reverse=True):
        if item["keyword"] in name:
            return item["status"], item["keyword"]
    return None, None


# ─── 분석 엔진 ──────────────────────────────────────────────────
def generate_report(brand_products, product_cats, product_f2_dates, today, resale_prevention, incoming_plan=None, product_status=None):
    results = []

    for code, info in brand_products.items():
        if code not in product_cats:
            continue
        if not info["lots"]:
            continue

        cat = product_cats[code]
        lots_sorted = sorted(info["lots"], key=lambda x: x["expiry"])
        total_stock = sum(l["stock"] for l in lots_sorted)

        f1_dates = [l["expiry"] for l in lots_sorted if l["stock"] > 0]
        f2_dates = product_f2_dates.get(code, [])

        if not f1_dates:
            continue

        # ── File1 우선 기준: min_expiry는 신재고입출고(File1) 기준 ──
        min_expiry = min(f1_dates)
        f2_min     = min(f2_dates) if f2_dates else None

        # 더블체크: File1·File2 최단 소비기한 차이 7일 초과 시 경고
        needs_doublecheck = bool(
            f2_min and abs((f2_min - min_expiry).days) > 7
        )
        f2_discrepancy = (
            f"File2({f2_min}) ≠ File1({min_expiry})" if needs_doublecheck else ""
        )

        # (하위 호환) f2_only_date 필드 유지 — 이제 경고 목적으로만 사용
        has_f2_only_date = False
        f2_only_min      = None

        months = months_remaining(min_expiry, today)

        # 카카오 제품: 온라인채널팀 담당
        is_kakao = "카카오" in info["name"]

        # 선물세트: 1개 = 2개월치 소비 → 3개월 전 조치 (일반: 4개월 전)
        is_gift_set = "선물세트" in info["name"]
        action_months = 3 if is_gift_set else 4
        raw_action  = min_expiry - relativedelta(months=action_months)
        action_date = adjust_to_weekday(raw_action)
        weekend_adj = action_date != raw_action

        if cat == "3개월":
            action    = "품절/단종 처리"
            cat_order = 0
        elif cat == "4-6개월":
            action    = "최상위 옵션 삭제 (4개월 이상 세트)"
            cat_order = 1
        else:
            action    = "최상위 옵션 삭제 (6개월 이상 세트)"
            cat_order = 2

        if action_date <= today:
            urgency       = "🔴 즉시 조치"
            urgency_order = 0
        elif action_date <= today + relativedelta(months=1):
            urgency       = "🟡 이번 달 조치"
            urgency_order = 1
        elif action_date <= today + relativedelta(months=3):
            urgency       = "🟠 2~3개월 내 조치"
            urgency_order = 2
        else:
            urgency       = "🟢 추후 확인"
            urgency_order = 3

        resale_info = resale_prevention.get(code)

        # 입고예정 확인 (3개월 내 조치 필요 항목만)
        incoming_info = None
        if incoming_plan is not None and urgency_order <= 2:
            plan = incoming_plan.get(code)
            if plan and plan["잔량"] > 0:
                incoming_info = {
                    "status": "입고예정",
                    "action": "품절처리 문의",
                    "잔량": plan["잔량"],
                    "예정일": plan["현재입고예정일"],
                }
            else:
                incoming_info = {
                    "status": "미입고예정",
                    "action": "단종처리 문의",
                    "잔량": 0,
                    "예정일": "",
                }

        results.append({
            "urgency_order":      urgency_order,
            "cat_order":          cat_order,
            "action_date":        action_date,
            "raw_action_date":    raw_action,
            "weekend_adjusted":   weekend_adj,
            "urgency":            urgency,
            "category":           cat,
            "brand":              info["brand"],
            "code":               code,
            "name":               info["name"],
            "min_expiry":         min_expiry,
            "months_left":        months,
            "total_stock":        total_stock,
            "action":             action,
            "lots":               lots_sorted,
            "has_f2_only_date":   has_f2_only_date,
            "f2_only_min":        f2_only_min,
            "has_resale":         resale_info is not None,
            "resale_info":        resale_info,
            "incoming_info":      incoming_info,
            "is_gift_set":        is_gift_set,
            "is_kakao":           is_kakao,
            "needs_doublecheck":  needs_doublecheck,
            "f2_discrepancy":     f2_discrepancy,
            "product_status":     match_product_status(info["name"], product_status or [])[0],
        })

    results.sort(key=lambda x: (x["urgency_order"], x["cat_order"], x["action_date"]))
    return results


def to_excel_bytes(results: list, today: date) -> bytes:
    rows = []
    for r in results:
        adj_note    = f" (원래 {r['raw_action_date']}, 주말→금요일 조정)" if r["weekend_adjusted"] else ""
        f2_note     = f" *파일2 보조확인({r['f2_only_min']})" if r["has_f2_only_date"] else ""
        resale_note = "✓ 재판매방지표 확인 필요" if r["has_resale"] else ""
        rows.append({
            "긴급도":         r["urgency"],
            "구분":           r["category"],
            "브랜드":         r["brand"],
            "상품코드":       r["code"],
            "품명":           r["name"],
            "최단 소비기한":  str(r["min_expiry"]) + f2_note,
            "남은 기간(개월)": r["months_left"],
            "총 재고(개)":    r["total_stock"],
            "액션 날짜":      str(r["action_date"]) + adj_note,
            "조치사항":       r["action"],
            "재판매방지":     resale_note,
        })
    df  = pd.DataFrame(rows)
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="임박재고 조치")
        ws = writer.sheets["임박재고 조치"]
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    return buf.getvalue()


# ─── 히스토리 ────────────────────────────────────────────────────
def load_history() -> dict:
    if "history_data" not in st.session_state:
        if os.path.exists(HISTORY_PATH):
            try:
                with open(HISTORY_PATH, "r", encoding="utf-8") as f:
                    st.session_state.history_data = json.load(f)
            except Exception:
                st.session_state.history_data = {}
        else:
            st.session_state.history_data = {}
    return st.session_state.history_data


def save_to_history(month_key, today_used, results) -> bool:
    history = load_history()
    if month_key in history:
        return False

    entry = {
        "date":  str(today_used),
        "month": month_key,
        "stats": {
            "total":      len(results),
            "urgent":     sum(1 for r in results if r["urgency_order"] == 0),
            "this_month": sum(1 for r in results if r["urgency_order"] == 1),
            "next_3m":    sum(1 for r in results if r["urgency_order"] == 2),
            "later":      sum(1 for r in results if r["urgency_order"] == 3),
        },
        "items": [
            {
                "urgency":          r["urgency"],
                "urgency_order":    r["urgency_order"],
                "category":         r["category"],
                "brand":            r["brand"],
                "code":             r["code"],
                "name":             r["name"],
                "min_expiry":       str(r["min_expiry"]),
                "months_left":      r["months_left"],
                "total_stock":      r["total_stock"],
                "action_date":      str(r["action_date"]),
                "action":           r["action"],
                "weekend_adjusted": r["weekend_adjusted"],
                "has_f2_only_date": r["has_f2_only_date"],
                "f2_only_min":      str(r["f2_only_min"]) if r["f2_only_min"] else None,
                "has_resale":       r["has_resale"],
            }
            for r in results
        ],
    }
    history[month_key] = entry
    st.session_state.history_data = history
    try:
        os.makedirs(os.path.dirname(HISTORY_PATH), exist_ok=True)
        with open(HISTORY_PATH, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    return True


# ─── 체크리스트 테이블 ──────────────────────────────────────────
def render_brand_checklist(results: list, brand: str, run_id: int):
    brand_items = [r for r in results if r["brand"] == brand]
    if not brand_items:
        st.info(f"이번 달 {brand} 대상 품목 없음")
        return

    rows = []
    for r in brand_items:
        resale_flag = "✌️" if r["has_resale"] else ""
        inc = r.get("incoming_info")
        if inc:
            inc_flag = "📦 입고예정" if inc["status"] == "입고예정" else "⛔ 미입고예정"
        else:
            inc_flag = ""
        gift_flag   = "🎁" if r.get("is_gift_set") else ""
        kakao_flag  = "🍫온라인채널" if r.get("is_kakao") else ""
        check_flag  = "⚠️더블체크" if r.get("needs_doublecheck") else ""
        rows.append({
            "✅ 완료": False,
            "재판매방지": resale_flag,
            "선물세트": gift_flag,
            "입고예정": inc_flag,
            "담당": kakao_flag,
            "더블체크": check_flag,
            "긴급도":     r["urgency"],
            "구분":       r["category"],
            "품명":       r["name"],
            "소비기한":   str(r["min_expiry"]),
            "남은 기간":  f"{r['months_left']}개월" if r["months_left"] is not None else "-",
            "재고(개)":   r["total_stock"],
            "액션 날짜":  str(r["action_date"]),
            "조치사항":   r["action"],
        })

    df = pd.DataFrame(rows)

    edited = st.data_editor(
        df,
        column_config={
            "✅ 완료":   st.column_config.CheckboxColumn("✅ 완료", default=False, width="small"),
            "재판매방지": st.column_config.TextColumn("✌️재판매", width="small"),
            "선물세트":  st.column_config.TextColumn("🎁세트", width="small"),
            "입고예정":  st.column_config.TextColumn("입고예정", width="small"),
            "담당":      st.column_config.TextColumn("담당", width="small"),
            "더블체크":  st.column_config.TextColumn("더블체크", width="small"),
            "긴급도":    st.column_config.TextColumn("긴급도", width="medium"),
            "구분":      st.column_config.TextColumn("구분", width="small"),
            "품명":      st.column_config.TextColumn("품명", width="large"),
            "소비기한":  st.column_config.TextColumn("소비기한", width="medium"),
            "남은 기간": st.column_config.TextColumn("남은 기간", width="small"),
            "재고(개)":  st.column_config.NumberColumn("재고(개)", format="%d개"),
            "액션 날짜": st.column_config.TextColumn("액션 날짜", width="medium"),
            "조치사항":  st.column_config.TextColumn("조치사항", width="large"),
        },
        disabled=["재판매방지","선물세트","입고예정","담당","더블체크","긴급도","구분","품명","소비기한","남은 기간","재고(개)","액션 날짜","조치사항"],
        hide_index=True,
        use_container_width=True,
        key=f"checklist_{brand}_{run_id}",
    )

    done  = int(edited["✅ 완료"].sum())
    total = len(brand_items)
    if total > 0:
        col_p, _ = st.columns([3, 2])
        with col_p:
            st.progress(done / total, text=f"완료 {done}/{total}개")

    # 재판매방지 항목 요약
    resale_items = [r for r in brand_items if r["has_resale"]]
    if resale_items:
        with st.expander(f"✌️ 재판매방지 확인 필요 항목 ({len(resale_items)}개)", expanded=False):
            for r in resale_items:
                ri = r["resale_info"]
                옵션목록 = ", ".join(ri.get("최소옵션", []))
                비고     = ri.get("비고", "")
                st.markdown(
                    f'<div class="resale-badge">'
                    f'<b>✌️ {r["name"]}</b> (코드: {r["code"]})<br>'
                    f'📌 재판매방지 출고 시작 옵션: <b>{옵션목록}</b><br>'
                    f'💡 추천 조치: 위 옵션부터 재판매방지 표에서 매핑된 SKU 유통기한·재고를 별도 확인 후 함께 조치'
                    + (f'<br>📝 비고: {비고}' if 비고 else '')
                    + '</div>',
                    unsafe_allow_html=True,
                )


# ─── 분석 페이지 ─────────────────────────────────────────────────
def page_analysis():
    # ── 사이드바 ──────────────────────────────────────────────
    with st.sidebar:
        st.markdown("##### 📂 파일1 — 신재고입출고")
        st.caption("YYMM_신재고입출고_그외팀사용**.xlsb**")
        file1_upload = st.file_uploader(
            "파일1", type=["xlsb"], label_visibility="collapsed", key="f1"
        )

        selected_team   = None
        selected_brands = []
        include_overseas = False

        if file1_upload:
            f1_bytes   = file1_upload.getvalue()
            stock_meta = read_stock_sheet(f1_bytes)

            if stock_meta["teams"]:
                st.markdown("##### 1차: 담당팀 선택 *(필수)*")
                selected_team = st.selectbox(
                    "팀", stock_meta["teams"], label_visibility="collapsed"
                )

                # 전체 팀에서 고유 브랜드 목록 수집 (순서 유지)
                all_brands: list = []
                seen_b: set = set()
                for _brands in stock_meta["team_brands"].values():
                    for _b in _brands:
                        if _b not in seen_b:
                            all_brands.append(_b)
                            seen_b.add(_b)

                st.markdown("##### 2차: 브랜드 선택 *(필수, 다중)*")
                selected_brands = st.multiselect(
                    "브랜드", all_brands,
                    default=None,
                    label_visibility="collapsed",
                    placeholder="브랜드를 선택하세요",
                )

                # 3차: 해외전용 포함 여부
                if selected_brands:
                    sel_codes: set = set()
                    for b in selected_brands:
                        for _team_codes in stock_meta["team_brand_codes"].values():
                            sel_codes.update(_team_codes.get(b, []))
                    has_overseas_in_sel = bool(sel_codes & stock_meta["overseas_codes"])
                    if has_overseas_in_sel:
                        st.markdown("##### 3차: 해외전용 포함 *(선택)*")
                        include_overseas = st.checkbox("해외전용 SKU 포함", value=False)
            else:
                st.warning("재고수불종합 시트를 찾을 수 없습니다.")
        else:
            f1_bytes   = None
            stock_meta = None

        st.markdown("##### 📂 파일2 — 유통기한 임박재고")
        st.caption("YY년 MM월 기준 유통기한 임박재고_공유용**.xlsx**")
        file2_upload = st.file_uploader(
            "파일2", type=["xlsx"], label_visibility="collapsed", key="f2"
        )

        st.divider()
        today_input = st.date_input("기준일", value=date.today())
        st.divider()

        can_run = bool(file1_upload and file2_upload and selected_brands)
        run_btn = st.button(
            "🔍 분석 실행", type="primary", use_container_width=True,
            disabled=not can_run,
        )
        if file1_upload and file2_upload and not selected_brands:
            st.caption("⚠️ 브랜드를 선택해야 실행됩니다")

        st.divider()
        with st.expander("사용 방법"):
            st.markdown("""
1. **파일1** `.xlsb` 업로드
2. 담당팀 → 브랜드 선택
3. (선택) 해외전용 포함 여부
4. **파일2** `.xlsx` 업로드
5. **분석 실행** 클릭
6. ✅ 조치 완료 표시
7. 히스토리 자동 저장
            """)

    # ── 메인 ─────────────────────────────────────────────────
    st.markdown("## 📦 유통기한 임박재고 분석 도구")
    st.caption("웰릿 · 클리너리 및 전 브랜드 | 팀 · 브랜드별 선택 분석 · 재판매방지 자동 감지")

    if not file1_upload or not file2_upload:
        st.info("👈 파일 2개를 업로드하고 브랜드를 선택한 뒤 **분석 실행**을 눌러주세요.")
        st.markdown("""
| 파일 | 설명 |
|------|------|
| 파일1 `.xlsb` | 신재고입출고 — 재고수불종합(팀/브랜드 선택) + 유통기한관리(로트/재고) |
| 파일2 `.xlsx` | 유통기한 임박재고 공유용 — 카테고리 · 소비기한 보조 확인 |
        """)
        st.stop()

    if not selected_brands:
        st.warning("사이드바에서 브랜드를 선택해 주세요.")
        st.stop()

    if run_btn:
        st.session_state.pop("results", None)
        st.session_state.pop("today_used", None)
        st.session_state.pop("history_saved_for", None)
        st.session_state["run_id"] = st.session_state.get("run_id", 0) + 1

    if run_btn or "results" in st.session_state:
        if "results" not in st.session_state:
            with st.spinner("파일 분석 중..."):
                try:
                    # 선택된 코드 집합 구성
                    sel_codes: set  = set()
                    code_to_brand: dict = {}
                    for brand in selected_brands:
                        for _tc in stock_meta["team_brand_codes"].values():
                          for c in _tc.get(brand, []):
                            if include_overseas or c not in stock_meta["overseas_codes"]:
                                sel_codes.add(c)
                                code_to_brand[c] = brand

                    code_whitelist       = tuple(sorted(sel_codes))
                    code_to_brand_tuple  = tuple(sorted(code_to_brand.items()))

                    f2_bytes = file2_upload.getvalue()
                    brand_products = load_file1(f1_bytes, code_whitelist, code_to_brand_tuple)
                    product_cats, product_f2_dates = load_file2(f2_bytes)
                    resale_prevention = load_resale_prevention()
                    incoming_plan = read_incoming_sheet(f1_bytes)
                    product_status = load_product_status()
                    results = generate_report(
                        brand_products, product_cats, product_f2_dates,
                        today_input, resale_prevention, incoming_plan, product_status,
                    )
                    st.session_state["results"]         = results
                    st.session_state["today_used"]      = today_input
                    st.session_state["selected_brands"] = selected_brands
                except Exception as e:
                    st.error(f"파일 처리 중 오류: {e}")
                    st.stop()

        results         = st.session_state["results"]
        today_used      = st.session_state.get("today_used", today_input)
        brands_display  = st.session_state.get("selected_brands", selected_brands)
        run_id          = st.session_state.get("run_id", 0)

        if not results:
            st.warning("매칭된 제품이 없습니다. 파일2에 대상 제품이 없거나 파일을 확인해 주세요.")
            st.stop()

        # ── 품절/단종 제품 분리 ─────────────────────────────
        excluded = [r for r in results if r.get("product_status") in ("품절", "단종")]
        results  = [r for r in results if r.get("product_status") not in ("품절", "단종")]

        if excluded:
            with st.expander(f"🚫 이미 품절/단종 처리된 제품 {len(excluded)}개 (분석 제외)", expanded=False):
                ex_rows = [{"상태": r["product_status"], "브랜드": r["brand"],
                            "품명": r["name"], "소비기한": str(r["min_expiry"])} for r in excluded]
                st.dataframe(pd.DataFrame(ex_rows), hide_index=True, use_container_width=True)

        # ── 히스토리 자동 저장 ──────────────────────────────
        month_key = today_used.strftime("%Y-%m")
        if st.session_state.get("history_saved_for") != month_key:
            newly = save_to_history(month_key, today_used, results)
            st.session_state["history_saved_for"] = month_key
            if newly:
                st.toast(f"✅ {month_key} 분석 결과 히스토리에 저장됨")

        # ── 재판매방지 경고 배너 ──────────────────────────────
        resale_count = sum(1 for r in results if r["has_resale"])
        if resale_count:
            st.warning(
                f"✌️ **재판매방지 확인 필요 품목 {resale_count}개** — "
                "아래 브랜드 탭에서 ✌️ 표시 품목을 확인하고 재판매방지표 매핑 SKU를 함께 조치하세요.",
                icon="⚠️",
            )

        # ── 요약 카드 ───────────────────────────────────────
        c1, c2, c3, c4, c5, c6 = st.columns(6)
        c1.metric("총 대상",    f"{len(results)}개")
        c2.metric("🔴 즉시",    f"{sum(1 for r in results if r['urgency_order']==0)}개")
        c3.metric("🟡 이번 달", f"{sum(1 for r in results if r['urgency_order']==1)}개")
        c4.metric("🟠 2~3개월", f"{sum(1 for r in results if r['urgency_order']==2)}개")
        c5.metric("🟢 추후 확인",f"{sum(1 for r in results if r['urgency_order']==3)}개")
        c6.metric("기준일",     str(today_used))

        st.divider()

        # ── 전체 조치 목록 (상단 / 브랜드 탭) ─────────────
        st.markdown("### 📋 전체 조치 목록")
        st.caption("✅ 완료 체크 | ✌️ = 재판매방지표 확인 필요")

        tab_labels = [
            f"{b} ({sum(1 for r in results if r['brand']==b)}개)"
            for b in brands_display
        ]
        brand_tabs = st.tabs(tab_labels)
        for tab, brand in zip(brand_tabs, brands_display):
            with tab:
                render_brand_checklist(results, brand, run_id)

        st.divider()

        # ── 엑셀 다운로드 ──────────────────────────────────
        col_dl, _ = st.columns([2, 5])
        with col_dl:
            st.download_button(
                "📥 리포트 다운로드 (Excel)",
                data=to_excel_bytes(results, today_used),
                file_name=f"유통기한임박재고_{today_used.strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        st.divider()

        # ── 상세 섹션 ────────────────────────────────────────
        SECTIONS = [
            {
                "title":    "이번 달 내 조치 항목",
                "subtitle": "🔴 즉시 조치 + 🟡 이번 달",
                "orders":   [0, 1],
                "expanded": True,
            },
            {
                "title":    "🟠 2~3개월 내 조치 예정",
                "subtitle": (
                    f"액션 날짜: {today_used + relativedelta(months=1) + timedelta(days=1)}"
                    f" ~ {today_used + relativedelta(months=3)}"
                ),
                "orders":   [2],
                "expanded": False,
            },
            {
                "title":    "🟢 추후 확인",
                "subtitle": f"액션 날짜: {today_used + relativedelta(months=3) + timedelta(days=1)} 이후",
                "orders":   [3],
                "expanded": False,
            },
        ]

        CAT_ACTION = {
            "3개월":   "품절/단종 처리",
            "4-6개월": "최상위 옵션 삭제",
            "7-9개월": "최상위 옵션 삭제",
        }

        def render_item(r, expanded):
            days_left = (r["action_date"] - today_used).days
            days_str  = (
                f"액션 {abs(days_left)}일 경과" if days_left < 0
                else "오늘 액션" if days_left == 0
                else f"D-{days_left}"
            )
            resale_tag = " ✌️" if r["has_resale"] else ""
            label = f"{r['urgency']}  |  **{r['brand']} {r['name']}**{resale_tag}  ({days_str})"

            with st.expander(label, expanded=expanded):
                col1, col2, col3, col4 = st.columns(4)
                col1.metric(
                    "최단 소비기한", str(r["min_expiry"]),
                    f"{r['months_left']}개월 남음" if r["months_left"] is not None else "",
                )
                col2.metric("총 재고", f"{r['total_stock']:,}개")
                col3.metric(
                    "액션 날짜", str(r["action_date"]),
                    f"주말→금 조정 (원래 {r['raw_action_date']})" if r["weekend_adjusted"] else None,
                )
                col4.metric("상품코드", r["code"])

                if r["has_f2_only_date"]:
                    st.markdown(
                        f'<div class="f2-note">📋 파일2 보조 확인 — '
                        f'파일1에 없는 소비기한 <b>{r["f2_only_min"]}</b>이 파일2에 있어 최단 소비기한으로 사용</div>',
                        unsafe_allow_html=True,
                    )

                if r.get("is_kakao"):
                    st.markdown(
                        '<div class="f2-note">🍫 <b>카카오 제품 — 온라인채널팀 담당</b>. '
                        '해당 팀과 협의 후 조치하세요.</div>',
                        unsafe_allow_html=True,
                    )

                if r.get("needs_doublecheck"):
                    st.markdown(
                        f'<div class="nostock-badge">⚠️ <b>소비기한 더블체크 필요</b> — '
                        f'신재고입출고(File1) 기준으로 분석했으나 파일 간 날짜 차이 발생<br>'
                        f'📋 {r["f2_discrepancy"]} — <b>신재고입출고 날짜가 최종 기준</b></div>',
                        unsafe_allow_html=True,
                    )

                if r.get("is_gift_set"):
                    st.markdown(
                        f'<div class="f2-note">🎁 <b>선물세트</b> — 1개 구매 시 2개월치 소비량. '
                        f'소비기한 <b>3개월 전</b> 조치 기준 적용 (일반 제품: 4개월 전)</div>',
                        unsafe_allow_html=True,
                    )

                # 재판매방지 경고
                if r["has_resale"]:
                    ri = r["resale_info"]
                    옵션목록 = ", ".join(ri.get("최소옵션", []))
                    비고     = ri.get("비고", "")
                    st.markdown(
                        f'<div class="resale-badge">'
                        f'✌️ <b>재판매방지표 확인 후 조치 필요</b><br>'
                        f'📌 재판매방지 출고 시작 옵션: <b>{옵션목록}</b><br>'
                        f'💡 추천 조치: 해당 옵션부터 재판매방지표에서 매핑된 SKU의 유통기한·재고를 별도 확인하고 일반 SKU와 함께 조치하세요.'
                        + (f'<br>📝 비고: {비고}' if 비고 else '')
                        + '</div>',
                        unsafe_allow_html=True,
                    )

                # 입고예정 여부 표시
                inc = r.get("incoming_info")
                if inc:
                    if inc["status"] == "입고예정":
                        예정일_txt = f" (입고예정일: {inc['예정일']})" if inc["예정일"] else ""
                        st.markdown(
                            f'<div class="incoming-badge">'
                            f'📦 <b>입고예정 확인됨{예정일_txt}</b> — 발주잔량 {inc["잔량"]:,}개<br>'
                            f'💡 추천 조치: <b>품절처리 문의</b> (신규 입고 예정이므로 단종 전 품절 처리)'
                            f'</div>',
                            unsafe_allow_html=True,
                        )
                    else:
                        st.markdown(
                            f'<div class="nostock-badge">'
                            f'⛔ <b>입고예정 없음</b> — 입고예정(구매) 시트에 해당 코드 없음<br>'
                            f'💡 추천 조치: <b>단종처리 문의</b> (추가 입고 계획 없으므로 단종 검토)'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

                badge_class = (
                    "urgent-badge"  if r["urgency_order"] == 0
                    else "warning-badge" if r["urgency_order"] == 1
                    else "orange-badge"  if r["urgency_order"] == 2
                    else "ok-badge"
                )
                prefix = (
                    "⚠️ 즉시 처리 필요 —" if r["urgency_order"] == 0
                    else "📅" if r["urgency_order"] == 1
                    else "🗓️" if r["urgency_order"] == 2
                    else "📌"
                )
                suffix = (
                    f"액션 날짜({r['action_date']}) 이미 지남." if r["urgency_order"] == 0
                    else f"{r['action_date']}까지 조치."
                    if r["urgency_order"] <= 2
                    else f"예정 액션: {r['action_date']}"
                )
                st.markdown(
                    f'<div class="{badge_class}">{prefix} <b>{r["action"]}</b> — {suffix}</div>',
                    unsafe_allow_html=True,
                )

                if len(r["lots"]) > 1:
                    lots_expiries = sorted(set(l["expiry"] for l in r["lots"]))
                    if len(lots_expiries) >= 2:
                        early, late = lots_expiries[0], lots_expiries[-1]
                        st.markdown(
                            f'<div class="incoming-badge">📦 <b>선입선출(FIFO) 안내</b> — '
                            f'유통기한 빠른 재고(<b>{early}</b>) 조치 시점에 '
                            f'유통기한 늦은 재고(<b>{late}</b>)로 교체 발송 준비</div>',
                            unsafe_allow_html=True,
                        )
                    st.caption(f"로트 상세 ({len(r['lots'])}개)")
                    lots_df = pd.DataFrame([
                        {
                            "소비기한": str(l["expiry"]),
                            "재고(개)": f"{l['stock']:,}",
                            "남은 기간": (
                                f"{months_remaining(l['expiry'], today_used)}개월"
                                if months_remaining(l["expiry"], today_used) is not None else "-"
                            ),
                        }
                        for l in r["lots"]
                    ])
                    st.dataframe(lots_df, hide_index=True, use_container_width=False)

        for sec in SECTIONS:
            sec_items = [r for r in results if r["urgency_order"] in sec["orders"]]
            if not sec_items:
                continue
            st.markdown(
                f"### {sec['title']} &nbsp;<span style='font-size:0.8em;color:#888;'>({len(sec_items)}개)</span>",
                unsafe_allow_html=True,
            )
            st.caption(sec["subtitle"])
            for cat_name in ["3개월", "4-6개월", "7-9개월"]:
                cat_items = [r for r in sec_items if r["category"] == cat_name]
                if not cat_items:
                    continue
                cat_icon = "🔴" if cat_name == "3개월" else "🟡" if cat_name == "4-6개월" else "🟢"
                st.markdown(
                    f"**{cat_icon} {cat_name} — {CAT_ACTION[cat_name]}** ({len(cat_items)}개)"
                )
                for r in cat_items:
                    render_item(r, expanded=sec["expanded"])
            st.divider()


# ─── 히스토리 페이지 ──────────────────────────────────────────────
def page_history():
    st.markdown("## 📋 분석 히스토리")
    st.caption("월별 분석 결과 이력 — 분석 실행 시 자동 저장 (같은 달 중복 없음)")

    history = load_history()

    col_import, col_export, _ = st.columns([2, 2, 3])
    with col_import:
        st.caption("이전 백업 파일 불러오기")
        uploaded_hist = st.file_uploader(
            "히스토리 JSON", type=["json"], key="hist_upload",
            label_visibility="collapsed",
        )
        if uploaded_hist:
            try:
                imported = json.load(uploaded_hist)
                existing = load_history()
                merged   = {**imported, **existing}
                st.session_state.history_data = merged
                try:
                    os.makedirs(os.path.dirname(HISTORY_PATH), exist_ok=True)
                    with open(HISTORY_PATH, "w", encoding="utf-8") as f:
                        json.dump(merged, f, ensure_ascii=False, indent=2)
                except Exception:
                    pass
                st.success(f"✅ 불러오기 완료 ({len(imported)}개월)")
                st.rerun()
            except Exception as e:
                st.error(f"불러오기 실패: {e}")

    with col_export:
        st.caption("히스토리 백업 저장")
        if history:
            hist_bytes = json.dumps(history, ensure_ascii=False, indent=2, default=str).encode("utf-8")
            st.download_button(
                "💾 히스토리 백업 (JSON)",
                data=hist_bytes,
                file_name="유통기한임박재고_히스토리.json",
                mime="application/json",
                use_container_width=True,
            )
        else:
            st.info("저장된 히스토리 없음")

    st.divider()

    if not history:
        st.info("아직 저장된 히스토리가 없습니다. 분석 실행 시 자동 저장됩니다.")
        return

    months_sorted = sorted(history.keys(), reverse=True)

    def items_to_df(item_list):
        return pd.DataFrame([
            {
                "긴급도":   i["urgency"],
                "구분":     i["category"],
                "품명":     ("✌️ " if i.get("has_resale") else "") + i["name"],
                "소비기한": i["min_expiry"],
                "남은 기간":f"{i['months_left']}개월" if i["months_left"] is not None else "-",
                "재고(개)": f"{i['total_stock']:,}",
                "액션 날짜":i["action_date"],
                "조치":     i["action"],
            }
            for i in item_list
        ])

    for month_key in months_sorted:
        entry = history[month_key]
        stats = entry.get("stats", {})
        items = entry.get("items", [])

        with st.expander(
            f"📅 {month_key}  —  총 {stats.get('total', 0)}개  (기준일: {entry.get('date', '?')})",
            expanded=(month_key == months_sorted[0]),
        ):
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("총 대상",   f"{stats.get('total', 0)}개")
            c2.metric("🔴 즉시",   f"{stats.get('urgent', 0)}개")
            c3.metric("🟡 이번 달",f"{stats.get('this_month', 0)}개")
            c4.metric("🟠+🟢 이후",f"{stats.get('next_3m', 0) + stats.get('later', 0)}개")

            if items:
                # 브랜드 목록 추출 (순서 유지)
                seen_brands: list = []
                for item in items:
                    if item["brand"] not in seen_brands:
                        seen_brands.append(item["brand"])

                tab_labels = [
                    f"{b} ({sum(1 for i in items if i['brand']==b)}개)"
                    for b in seen_brands
                ]
                tabs = st.tabs(tab_labels)
                for tab, brand in zip(tabs, seen_brands):
                    with tab:
                        brand_items = [i for i in items if i["brand"] == brand]
                        if brand_items:
                            st.dataframe(items_to_df(brand_items), hide_index=True, use_container_width=True)
                        else:
                            st.info(f"{brand} 없음")


# ─── 품절/단종 관리 페이지 ───────────────────────────────────────
def page_status():
    st.markdown("## 🚫 품절/단종 제품 관리")
    st.caption("이 목록에 등록된 제품은 분석에서 자동 제외됩니다.")

    status_list = load_product_status()

    # ── 현재 목록 ──
    if status_list:
        st.markdown("### 현재 목록")
        df = pd.DataFrame(status_list)[["brand", "keyword", "status"]]
        df.columns = ["브랜드", "키워드", "상태"]
        st.dataframe(df, hide_index=True, use_container_width=True)
    else:
        st.info("등록된 품절/단종 제품 없음")

    st.divider()

    # ── 추가 ──
    st.markdown("### ➕ 새 제품 추가")
    col1, col2, col3 = st.columns(3)
    new_kw     = col1.text_input("품명 키워드", placeholder="예: 이뮨베라 키즈")
    new_brand  = col2.selectbox("브랜드", ["웰릿", "클리너리", "기타"])
    new_status = col3.selectbox("상태", ["품절", "단종"])

    if st.button("추가", type="primary", disabled=not new_kw.strip()):
        if any(i["keyword"] == new_kw.strip() for i in status_list):
            st.warning("이미 등록된 키워드입니다.")
        else:
            status_list.append({"keyword": new_kw.strip(), "brand": new_brand, "status": new_status})
            save_product_status(status_list)
            st.success(f"✅ '{new_kw.strip()}' 추가됨")
            st.rerun()

    st.divider()

    # ── 삭제 ──
    if status_list:
        st.markdown("### 🗑️ 제품 삭제")
        del_kw = st.selectbox(
            "삭제할 항목 선택",
            options=[i["keyword"] for i in status_list],
        )
        if st.button("삭제", type="secondary"):
            status_list = [i for i in status_list if i["keyword"] != del_kw]
            save_product_status(status_list)
            st.success(f"🗑️ '{del_kw}' 삭제됨")
            st.rerun()

    st.divider()
    st.caption("변경 후 GitHub에 push해야 배포 앱에 반영됩니다.")


# ─── 메인 라우터 ──────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📦 유통기한 임박재고")
    page = st.radio(
        "페이지", ["📊 분석", "📋 히스토리", "🚫 품절/단종 관리"],
        label_visibility="collapsed",
    )
    st.divider()

if page == "📊 분석":
    page_analysis()
elif page == "📋 히스토리":
    page_history()
elif page == "🚫 품절/단종 관리":
    page_status()
